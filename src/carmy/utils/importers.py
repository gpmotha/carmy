"""Data importers for CSV, JSON, XLSX files, and historical SQLite databases."""

import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from carmy.models.meal import Cuisine, Meal, MealIngredient, MealType
from carmy.models.plan import MealSlot, PlanMeal, WeeklyPlan


def normalize_meal_type(raw_type: str | None) -> str:
    """Normalize meal type to enum value."""
    if not raw_type:
        return MealType.MAIN_COURSE.value

    type_mapping = {
        "soup": MealType.SOUP.value,
        "main course": MealType.MAIN_COURSE.value,
        "pasta": MealType.PASTA.value,
        "salad": MealType.SALAD.value,
        "dessert": MealType.DESSERT.value,
        "breakfast": MealType.BREAKFAST.value,
        "appetizer": MealType.APPETIZER.value,
        "spread": MealType.SPREAD.value,
        "beverage": MealType.BEVERAGE.value,
        "condiment": MealType.CONDIMENT.value,
        "dinner": MealType.DINNER.value,
        "meat": MealType.MEAT.value,
        "pastry": MealType.PASTRY.value,
        "vegan": MealType.VEGAN.value,
    }

    normalized = raw_type.strip().lower()
    return type_mapping.get(normalized, MealType.MAIN_COURSE.value)


def normalize_cuisine(raw_cuisine: str | None) -> str | None:
    """Normalize cuisine to enum value."""
    if not raw_cuisine:
        return None

    cuisine_mapping = {
        "hungarian": Cuisine.HUNGARIAN.value,
        "italian": Cuisine.ITALIAN.value,
        "french": Cuisine.FRENCH.value,
        "indian": Cuisine.INDIAN.value,
        "middle eastern": Cuisine.MIDDLE_EASTERN.value,
        "american": Cuisine.AMERICAN.value,
        "asian": Cuisine.ASIAN.value,
        "international": Cuisine.INTERNATIONAL.value,
        "austrian": Cuisine.AUSTRIAN.value,
        "balkan": Cuisine.BALKAN.value,
        "belgian": Cuisine.BELGIAN.value,
        "british": Cuisine.BRITISH.value,
        "cuban": Cuisine.CUBAN.value,
        "german": Cuisine.GERMAN.value,
        "greek": Cuisine.GREEK.value,
        "mexican": Cuisine.MEXICAN.value,
    }

    normalized = raw_cuisine.strip().lower()
    return cuisine_mapping.get(normalized, Cuisine.INTERNATIONAL.value)


def get_iso_week(dt: date | datetime) -> tuple[int, int]:
    """Get ISO year and week number from a date."""
    if isinstance(dt, datetime):
        dt = dt.date()
    iso_cal = dt.isocalendar()
    return iso_cal[0], iso_cal[1]


def get_week_start(year: int, week: int) -> date:
    """Get the Monday of a given ISO week."""
    jan_4 = date(year, 1, 4)
    start_of_week_1 = jan_4 - __import__("datetime").timedelta(days=jan_4.weekday())
    return start_of_week_1 + __import__("datetime").timedelta(weeks=week - 1)


class XLSXImporter:
    """Import historical meal plan data from XLSX files."""

    def __init__(self, session: Session):
        self.session = session
        self.meals_cache: dict[str, Meal] = {}
        self.plans_cache: dict[tuple[int, int], WeeklyPlan] = {}
        self.stats = {
            "meals_created": 0,
            "meals_updated": 0,
            "plans_created": 0,
            "plan_meals_created": 0,
            "skipped": 0,
        }

    def import_history(self, file_path: Path | str) -> dict[str, int]:
        """Import historical data from Meal_planner.xlsx.

        The file is expected to have:
        - Plan sheet: Week, Start, Név, Name, Type, Cuisine, Category, etc.
        - Sheet2: Unique meal catalog (Name, Type, Cuisine)

        Args:
            file_path: Path to the XLSX file

        Returns:
            Dict with import statistics
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(file_path, data_only=True)

        # First, build meal catalog from Sheet2 (cleaner unique list)
        if "Sheet2" in wb.sheetnames:
            self._import_meal_catalog(wb["Sheet2"])

        # Then import historical plans from Plan sheet
        if "Plan" in wb.sheetnames:
            self._import_plan_history(wb["Plan"])

        self.session.commit()
        return self.stats

    def _import_meal_catalog(self, ws) -> None:
        """Import unique meals from Sheet2."""
        # Headers: Name, Type, Cuisine
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            name = row[0].value
            meal_type = row[1].value
            cuisine = row[2].value

            if not name:
                continue

            self._get_or_create_meal(
                nev=name,  # Will use English name as Hungarian if not found
                name=name,
                meal_type=meal_type,
                cuisine=cuisine,
            )

    def _import_plan_history(self, ws) -> None:
        """Import historical weekly plans from Plan sheet.

        Headers: Week, Start, Név, Name, Type, Cuisine, Category, Calories,
                 Prep time, Cook time, Difficulty, Seasonality, Ingredients, Note,
                 Default Portions, Keeps Days
        """
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            week_num = row[0].value
            start_date = row[1].value
            nev = row[2].value
            name = row[3].value
            meal_type = row[4].value
            cuisine = row[5].value
            category = row[6].value
            calories = row[7].value
            prep_time = row[8].value
            cook_time = row[9].value
            difficulty = row[10].value
            seasonality = row[11].value
            ingredients = row[12].value
            note = row[13].value
            # New fields (columns O and P, indices 14 and 15)
            default_portions = row[14].value if len(row) > 14 else None
            keeps_days = row[15].value if len(row) > 15 else None

            if not name or not week_num:
                self.stats["skipped"] += 1
                continue

            # Get or create the meal
            meal = self._get_or_create_meal(
                nev=nev or name,
                name=name,
                meal_type=meal_type,
                cuisine=cuisine,
                category=category,
                calories=calories,
                prep_time=prep_time,
                cook_time=cook_time,
                difficulty=difficulty,
                seasonality=seasonality,
                ingredients=ingredients,
                default_portions=default_portions,
                keeps_days=keeps_days,
            )

            # Get or create the weekly plan
            if isinstance(start_date, datetime):
                year = start_date.year
                plan_start = start_date.date()
            else:
                year = start_date.year if start_date else 2023
                plan_start = start_date if isinstance(start_date, date) else get_week_start(
                    year, int(week_num)
                )

            plan = self._get_or_create_plan(year, int(week_num), plan_start)

            # Check if this meal is already in this plan
            existing = (
                self.session.query(PlanMeal)
                .filter(PlanMeal.plan_id == plan.id, PlanMeal.meal_id == meal.id)
                .first()
            )

            if not existing:
                is_leftover = bool(note and "leftover" in note.lower()) if note else False
                plan_meal = PlanMeal(
                    plan_id=plan.id,
                    meal_id=meal.id,
                    is_leftover=is_leftover,
                    notes=note,
                )
                self.session.add(plan_meal)
                self.stats["plan_meals_created"] += 1

    def _get_or_create_meal(
        self,
        nev: str,
        name: str,
        meal_type: str | None = None,
        cuisine: str | None = None,
        category: str | None = None,
        calories: int | None = None,
        prep_time: int | None = None,
        cook_time: int | None = None,
        difficulty: str | None = None,
        seasonality: str | None = None,
        ingredients: str | None = None,
        default_portions: int | None = None,
        keeps_days: int | None = None,
    ) -> Meal:
        """Get existing meal or create new one."""
        cache_key = name.lower().strip()

        if cache_key in self.meals_cache:
            return self.meals_cache[cache_key]

        # Try to find existing meal
        existing = self.session.query(Meal).filter(Meal.name.ilike(name.strip())).first()

        if existing:
            self.meals_cache[cache_key] = existing
            # Update with new info if available
            if calories and not existing.calories:
                existing.calories = int(calories) if calories else None
            if prep_time and not existing.prep_time_minutes:
                existing.prep_time_minutes = int(prep_time) if prep_time else 0
            if cook_time and not existing.cook_time_minutes:
                existing.cook_time_minutes = int(cook_time) if cook_time else 0
            if default_portions and existing.default_portions == 1:
                existing.default_portions = int(default_portions)
            if keeps_days and existing.keeps_days == 1:
                existing.keeps_days = int(keeps_days)
            self.stats["meals_updated"] += 1
            return existing

        # Create new meal
        normalized_type = normalize_meal_type(meal_type)
        normalized_cuisine = normalize_cuisine(cuisine)

        # Determine meat/vegetarian flags from type
        has_meat = normalized_type in [MealType.MEAT.value]
        is_vegan = normalized_type in [MealType.VEGAN.value]
        is_vegetarian = is_vegan or normalized_type not in [
            MealType.MEAT.value,
            MealType.MAIN_COURSE.value,
        ]

        meal = Meal(
            nev=nev.strip() if nev else name.strip(),
            name=name.strip(),
            meal_type=normalized_type,
            cuisine=normalized_cuisine,
            category=category.lower().strip() if category else None,
            calories=int(calories) if calories else None,
            prep_time_minutes=int(prep_time) if prep_time else 0,
            cook_time_minutes=int(cook_time) if cook_time else 0,
            difficulty=(difficulty.lower().strip() if difficulty else "easy"),
            seasonality=(seasonality.lower().strip() if seasonality else "year_round"),
            is_vegetarian=is_vegetarian,
            is_vegan=is_vegan,
            has_meat=has_meat,
            default_portions=int(default_portions) if default_portions else 1,
            keeps_days=int(keeps_days) if keeps_days else 1,
        )
        self.session.add(meal)
        self.session.flush()  # Get the ID

        # Add ingredients if provided
        if ingredients:
            for ing in ingredients.split(","):
                ing = ing.strip()
                if ing:
                    meal_ing = MealIngredient(
                        meal_id=meal.id,
                        ingredient=ing.lower(),
                        is_flavor_base=False,  # Could be enhanced later
                    )
                    self.session.add(meal_ing)

        self.meals_cache[cache_key] = meal
        self.stats["meals_created"] += 1
        return meal

    def _get_or_create_plan(self, year: int, week: int, start_date: date) -> WeeklyPlan:
        """Get existing weekly plan or create new one."""
        cache_key = (year, week)

        if cache_key in self.plans_cache:
            return self.plans_cache[cache_key]

        existing = (
            self.session.query(WeeklyPlan)
            .filter(WeeklyPlan.year == year, WeeklyPlan.week_number == week)
            .first()
        )

        if existing:
            self.plans_cache[cache_key] = existing
            return existing

        plan = WeeklyPlan(
            year=year,
            week_number=week,
            start_date=start_date,
        )
        self.session.add(plan)
        self.session.flush()

        self.plans_cache[cache_key] = plan
        self.stats["plans_created"] += 1
        return plan


class HistoricalDBImporter:
    """Import historical meal data from carmy_historical.db SQLite database."""

    # Map Hungarian meal time to MealSlot
    MEAL_TIME_MAP = {
        "ebéd": MealSlot.LUNCH.value,
        "reggeli": MealSlot.BREAKFAST.value,
        "vacsora": MealSlot.DINNER.value,
        "uzsonna": MealSlot.SNACK.value,
        "desszert": MealSlot.SNACK.value,
    }

    # Map Hungarian day names to day_of_week (0=Monday)
    DAY_MAP = {
        "hétfő": 0,
        "kedd": 1,
        "szerda": 2,
        "csütörtök": 3,
        "péntek": 4,
        "szombat": 5,
        "vasárnap": 6,
    }

    def __init__(self, session: Session):
        self.session = session
        self.meals_cache: dict[str, Meal] = {}
        self.plans_cache: dict[tuple[int, int], WeeklyPlan] = {}
        self.stats = {
            "meals_created": 0,
            "meals_skipped": 0,
            "plans_created": 0,
            "plans_skipped": 0,
            "plan_meals_created": 0,
            "plan_meals_skipped": 0,
        }

    def import_from_db(self, db_path: Path | str) -> dict[str, int]:
        """Import historical data from carmy_historical.db.

        Args:
            db_path: Path to the SQLite database file

        Returns:
            Dict with import statistics
        """
        db_path = Path(db_path)
        if not db_path.exists():
            raise FileNotFoundError(f"Database not found: {db_path}")

        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        try:
            # Step 1: Import unique meals to catalog
            self._import_meals(cursor)

            # Step 2: Import weekly plans
            self._import_weekly_plans(cursor)

            # Step 3: Import daily meals and link to plans
            self._import_daily_meals(cursor)

            self.session.commit()
        finally:
            conn.close()

        return self.stats

    def _import_meals(self, cursor: sqlite3.Cursor) -> None:
        """Import unique meals from daily_meal table to Carmy's meal catalog."""
        # Get unique meal names from historical data
        cursor.execute("""
            SELECT DISTINCT meal_name_hu, meal_name_en
            FROM daily_meal
            WHERE meal_name_hu IS NOT NULL AND meal_name_hu != ''
            ORDER BY meal_name_hu
        """)

        for row in cursor.fetchall():
            name_hu = row["meal_name_hu"].strip()
            name_en = row["meal_name_en"].strip() if row["meal_name_en"] else None

            # Skip empty names
            if not name_hu:
                continue

            # Check if meal already exists
            cache_key = name_hu.lower()
            if cache_key in self.meals_cache:
                continue

            existing = (
                self.session.query(Meal)
                .filter(Meal.nev.ilike(name_hu))
                .first()
            )

            if existing:
                self.meals_cache[cache_key] = existing
                self.stats["meals_skipped"] += 1
                continue

            # Infer meal type from name
            meal_type = self._infer_meal_type(name_hu)

            # Create new meal
            meal = Meal(
                nev=name_hu,
                name=name_en or self._transliterate_name(name_hu),
                meal_type=meal_type,
                cuisine=Cuisine.HUNGARIAN.value,  # Default to Hungarian
                difficulty="easy",
                seasonality="year_round",
            )
            self.session.add(meal)
            self.session.flush()

            self.meals_cache[cache_key] = meal
            self.stats["meals_created"] += 1

    def _import_weekly_plans(self, cursor: sqlite3.Cursor) -> None:
        """Import weekly plans from historical database."""
        cursor.execute("""
            SELECT id, year, week_number, start_date, end_date, source, notes
            FROM weekly_plan
            ORDER BY year, week_number
        """)

        for row in cursor.fetchall():
            year = row["year"]
            week = row["week_number"]
            cache_key = (year, week)

            if cache_key in self.plans_cache:
                continue

            # Check if plan already exists in Carmy
            existing = (
                self.session.query(WeeklyPlan)
                .filter(WeeklyPlan.year == year, WeeklyPlan.week_number == week)
                .first()
            )

            if existing:
                self.plans_cache[cache_key] = existing
                self.stats["plans_skipped"] += 1
                continue

            # Parse date
            start_date = datetime.strptime(row["start_date"], "%Y-%m-%d").date()

            # Create new plan
            plan = WeeklyPlan(
                year=year,
                week_number=week,
                start_date=start_date,
                notes=row["notes"],
            )
            self.session.add(plan)
            self.session.flush()

            self.plans_cache[cache_key] = plan
            self.stats["plans_created"] += 1

    def _import_daily_meals(self, cursor: sqlite3.Cursor) -> None:
        """Import daily meal assignments."""
        cursor.execute("""
            SELECT dm.*, wp.year, wp.week_number
            FROM daily_meal dm
            JOIN weekly_plan wp ON dm.plan_id = wp.id
            ORDER BY wp.year, wp.week_number, dm.meal_date
        """)

        for row in cursor.fetchall():
            year = row["year"]
            week = row["week_number"]
            name_hu = row["meal_name_hu"]

            if not name_hu:
                continue

            cache_key = (year, week)
            plan = self.plans_cache.get(cache_key)
            if not plan:
                self.stats["plan_meals_skipped"] += 1
                continue

            meal_cache_key = name_hu.lower().strip()
            meal = self.meals_cache.get(meal_cache_key)
            if not meal:
                # Try to find by similar name
                meal = (
                    self.session.query(Meal)
                    .filter(Meal.nev.ilike(f"%{name_hu}%"))
                    .first()
                )
                if not meal:
                    self.stats["plan_meals_skipped"] += 1
                    continue

            # Parse meal date for day_of_week
            meal_date = datetime.strptime(row["meal_date"], "%Y-%m-%d").date()
            day_of_week = meal_date.weekday()  # 0=Monday

            # Map meal_time
            meal_time = row["meal_time"] or "ebéd"
            meal_slot = self.MEAL_TIME_MAP.get(meal_time, MealSlot.LUNCH.value)

            # Check if already exists
            existing = (
                self.session.query(PlanMeal)
                .filter(
                    PlanMeal.plan_id == plan.id,
                    PlanMeal.meal_id == meal.id,
                    PlanMeal.day_of_week == day_of_week,
                    PlanMeal.meal_slot == meal_slot,
                )
                .first()
            )

            if existing:
                self.stats["plan_meals_skipped"] += 1
                continue

            # Create plan meal
            is_leftover = bool(row["is_leftover"])
            plan_meal = PlanMeal(
                plan_id=plan.id,
                meal_id=meal.id,
                day_of_week=day_of_week,
                meal_slot=meal_slot,
                is_leftover=is_leftover,
                notes=row["notes"],
            )
            self.session.add(plan_meal)
            self.stats["plan_meals_created"] += 1

    def _infer_meal_type(self, name_hu: str) -> str:
        """Infer meal type from Hungarian name."""
        name_lower = name_hu.lower()

        # Soup patterns
        if "leves" in name_lower:
            return MealType.SOUP.value

        # Pasta patterns
        if any(x in name_lower for x in ["tészta", "pasta", "lasagne", "rigatoni", "carbonara", "ravioli", "tortellini"]):
            return MealType.PASTA.value

        # Dessert patterns
        if any(x in name_lower for x in ["torta", "süti", "gombóc", "galuska", "palacsinta", "kuglóf", "kalács", "muffin"]):
            return MealType.DESSERT.value

        # Salad patterns
        if "saláta" in name_lower:
            return MealType.SALAD.value

        # Breakfast patterns
        if any(x in name_lower for x in ["reggeli", "tojás", "bacon", "pirítós", "tejbegríz"]):
            return MealType.BREAKFAST.value

        # Appetizer/spread patterns
        if any(x in name_lower for x in ["krém", "humusz", "padlizsán"]):
            return MealType.APPETIZER.value

        # Default to main course
        return MealType.MAIN_COURSE.value

    def _transliterate_name(self, name_hu: str) -> str:
        """Convert Hungarian name to English-friendly version."""
        # Simple transliteration for common characters
        replacements = {
            "á": "a", "é": "e", "í": "i", "ó": "o", "ö": "o", "ő": "o",
            "ú": "u", "ü": "u", "ű": "u",
            "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ö": "O", "Ő": "O",
            "Ú": "U", "Ü": "U", "Ű": "U",
        }
        result = name_hu
        for hun, eng in replacements.items():
            result = result.replace(hun, eng)
        return result
